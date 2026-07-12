import os
import re
import shutil
import pandas as pd
import requests
from openpyxl import load_workbook

MASTER_SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1qEzCT6bvpGEm2bjf0TR0WK-iJ2yMyiE_Ol7BuLID1Gs/export?format=xlsx"
OUTPUT_DIR = "./bronze_data"

def sanitize_filename(name):
    name = str(name).strip()
    # Replace illegal characters in directory/file names with underscore
    name = re.sub(r'[\\/*?:"<>|]', '_', name)
    return name

def get_spreadsheet_id(url):
    if not isinstance(url, str):
        return None
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    return match.group(1) if match else None

def to_export_url(url):
    spreadsheet_id = get_spreadsheet_id(url)
    if not spreadsheet_id:
        return None
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"

def fetch_spreadsheet_title(url):
    try:
        # Try to extract the title from the spreadsheet edit HTML
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        res = requests.get(url, headers=headers, timeout=10)
        if res.status_code == 200:
            match = re.search(r'<title>(.*?)</title>', res.text)
            if match:
                title = match.group(1)
                if title.endswith(" - Google Sheets"):
                    title = title[:-16]
                return title.strip()
    except Exception as e:
        print(f"   ⚠️ Warning: Failed to fetch online title for {url} ({e})")
    return None

def main():
    print(f"🧹 Preparing output directory '{OUTPUT_DIR}'...")
    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"📥 Downloading master spreadsheet ledger...")
    master_scratch = "master_ledger.xlsx"
    try:
        res = requests.get(MASTER_SPREADSHEET_URL, timeout=15)
        if res.status_code != 200:
            print(f"❌ Failed to download master ledger (HTTP {res.status_code})")
            return
        with open(master_scratch, "wb") as f:
            f.write(res.content)
    except Exception as e:
        print(f"❌ Error downloading master ledger: {e}")
        return

    try:
        # Load the MAIN sheet from the master ledger workbook
        xl_master = pd.ExcelFile(master_scratch)
        if 'MAIN' not in xl_master.sheet_names:
            print("❌ The master ledger does not contain a 'MAIN' sheet.")
            return

        df = pd.read_excel(xl_master, sheet_name='MAIN')
    except Exception as e:
        print(f"❌ Error reading master ledger: {e}")
        return
    finally:
        if os.path.exists(master_scratch):
            os.remove(master_scratch)

    # Ensure required columns are present
    required_cols = ['season_id', 'game_id', 'division', 'data_type', 'access']
    for col in required_cols:
        if col not in df.columns:
            print(f"❌ Missing required column '{col}' in MAIN sheet.")
            return

    scratchpad = "raw_download.xlsx"

    print(f"🚀 Starting dynamic bronze data extraction...")

    # Pass 1: resolve every unique spreadsheet to a directory name up front.
    # Distinct spreadsheets can share a title (e.g. two "TFT Scoreboard"
    # sheets); every member of a colliding title group gets a short-id
    # suffix, so the name mapping is stable regardless of ledger row order.
    sheets = {}  # spreadsheet_id -> {'url': ..., 'dir_name': ...}
    for idx, row in df.iterrows():
        url = row['access']
        spreadsheet_id = get_spreadsheet_id(url)

        # Skip if no valid spreadsheet ID found; de-duplicate repeat URLs
        if not spreadsheet_id or spreadsheet_id in sheets:
            continue

        # Fallback name constructed from metadata: season_id, game_id, division, data_type
        meta_parts = []
        for key in ['season_id', 'game_id', 'division', 'data_type']:
            val = row[key]
            if pd.notna(val) and str(val).strip() and str(val).lower() != 'nan':
                meta_parts.append(str(val).strip())
        fallback_name = "_".join(meta_parts) if meta_parts else f"spreadsheet_{spreadsheet_id}"

        title = fetch_spreadsheet_title(url)
        if title:
            print(f"   🏷️ Row {idx+1}: '{title}'")
        else:
            print(f"   ⚠️ Row {idx+1}: using fallback name '{fallback_name}'")
        sheets[spreadsheet_id] = {'url': url, 'dir_name': sanitize_filename(title or fallback_name)}

    ids_by_name = {}
    for spreadsheet_id, info in sheets.items():
        ids_by_name.setdefault(info['dir_name'], []).append(spreadsheet_id)
    for dir_name, ids in ids_by_name.items():
        if len(ids) > 1:
            print(f"⚠️ Title '{dir_name}' is shared by {len(ids)} spreadsheets; suffixing each with its id.")
            for spreadsheet_id in ids:
                sheets[spreadsheet_id]['dir_name'] = f"{dir_name}_{spreadsheet_id[:6]}"

    # Pass 2: download and dump every spreadsheet.
    for spreadsheet_id, info in sheets.items():
        url, dir_name = info['url'], info['dir_name']
        print(f"\n📂 Processing spreadsheet ID {spreadsheet_id} -> '{dir_name}'...")

        sub_dir_path = os.path.join(OUTPUT_DIR, dir_name)
        os.makedirs(sub_dir_path, exist_ok=True)

        # 2. Download spreadsheet
        dl_url = to_export_url(url)
        print(f"   📥 Downloading workbook...")
        try:
            res = requests.get(dl_url, timeout=20)
            if res.status_code != 200:
                print(f"   ⚠️ Failed to download (HTTP {res.status_code})")
                continue

            with open(scratchpad, "wb") as f:
                f.write(res.content)

            # 3. Read and export all sheets
            xl = pd.ExcelFile(scratchpad)
            for sheet_name in xl.sheet_names:
                df_sheet = pd.read_excel(scratchpad, sheet_name=sheet_name, header=None)

                safe_sheet_name = sanitize_filename(sheet_name)
                csv_path = os.path.join(sub_dir_path, f"{safe_sheet_name}.csv")

                df_sheet.to_csv(csv_path, index=False, header=False)
                print(f"   💾 Saved sheet '{sheet_name}' to: {csv_path}")

            # 4. Export a bold mask per sheet — some schedules mark the match
            # winner only via bold formatting, which the CSV dump loses.
            wb = load_workbook(scratchpad)
            for ws in wb.worksheets:
                mask = [['1' if (c.font and c.font.bold) else '0' for c in row] for row in ws.iter_rows()]
                if not mask:
                    continue
                bold_path = os.path.join(sub_dir_path, f"{sanitize_filename(ws.title)}__bold.csv")
                with open(bold_path, 'w') as f:
                    f.write('\n'.join(','.join(r) for r in mask))

        except Exception as e:
            print(f"   ❌ Error extracting spreadsheet: {e}")

    if os.path.exists(scratchpad):
        os.remove(scratchpad)

    print(f"\n🎉 Finished! All bronze data is organized hierarchically in '{OUTPUT_DIR}'.")

if __name__ == "__main__":
    main()
